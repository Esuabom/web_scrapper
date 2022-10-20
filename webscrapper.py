#Importing libraries
from multiprocessing.connection import wait
from openpyxl import workbook, load_workbook
from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC 
from selenium.webdriver.support.wait import WebDriverWait 
import time 

#Reading excel file 
wbpath = "C:/Users/Theo Hernandez/Desktop/Massasuchettes Database.xlsx"
wb = load_workbook(wbpath)
ws = wb["Sheet1"]

#Activating selenium webdriver
PATH = "C:/Program Files (x86)/chromedriver_win32/chromedriver.exe"
driver = webdriver.Chrome(PATH)
    
#Navigating to url
url = "insert your url"
driver.get(url)
try:
        #defining element variables
        aaname = "Hosting Capacity"
        class_name1 = "checkbox jimu-float-leading jimu-icon jimu-icon-checkbox"
        
        element = WebDriverWait(driver, 10). until(EC.presence_of_element_located((By.ID, aaname)))
        driver.find_element(By.NAME, aaname).click()
except:
        driver.quit()
    
try:     
        element1= WebDriverWait(driver, 10). until(EC.presence_of_element_located((By.CLASS_NAME, class_name1)))
        driver.find_element(By.CLASS_NAME, class_name1).click()
except:
        driver.quit()
try:
        id_element = "esri_dijit_Search_0_input"
        element2= WebDriverWait(driver, 10). until(EC.presence_of_element_located((By.ID, id_element)))   
        
except:
        driver.quit()

#Iterating through excel worksheet and passing data into webpage
for row in ws.iter_rows (min_row=2, values_only= True):
    search_query = str(row[1]) + ", " + str(row[2]) + ", " + str(row[3])
    print (search_query)

    search = driver.find_element(By.ID, id_element)
    search.send_keys(search_query)
    search.send_keys(Keys.RETURN)

    try:
        class_name2 =  "esriPopup esriPopupVisible"
        element3 = WebDriverWait(driver, 15). until (EC.presence_of_element_located((By.CLASS_NAME, class_name2)))
    except:
        driver.quit()
    
    visibility = driver.find_element(By.CLASS_NAME, class_name2)
    visibility.click()
    time.sleep(5)

    #passing scraped data back to excel file
    data_retrieved = driver.find_element(By.TAG_NAME, "tab").text
    print(data_retrieved)
    row[4] = str(data_retrieved)

wb.save(wbpath)
#end of code

